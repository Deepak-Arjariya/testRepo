import boto3
import openpyxl

# Load the XLSX file
workbook = openpyxl.load_workbook('data.xlsx')
sheet = workbook.active

# Get header row for column mapping
header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
header_mapping = {value: index for index, value in enumerate(header_row)}

# Initialize DynamoDB client
dynamodb = boto3.client('dynamodb', region_name='your_region_name')

# Iterate through the rows in the XLSX file
for row in sheet.iter_rows(min_row=2, values_only=True):
    partition_key = row[header_mapping['PartitionKey']]  # Replace 'PartitionKey' with your actual header name

    # Retrieve existing item from DynamoDB
    response = dynamodb.get_item(
        TableName='your_table_name',
        Key={'PartitionKeyName': {'S': partition_key}}
    )

    existing_item = response.get('Item', {})

    # Merge data from Excel sheet with existing item, overwriting existing values
    for header, update_value in row[1:]:
        col_name = header  # Assuming header is a valid attribute name

        if update_value:
            existing_item[col_name] = {'S': update_value}

    # Perform update request
    response = dynamodb.put_item(
        TableName='your_table_name',
        Item=existing_item
    )

    print(f"Updated item with partition key: {partition_key}")
    
